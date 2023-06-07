import pandas as pd
from openpyxl import load_workbook

# Read the Excel file
file_path = 'loss_and_accurate.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# Convert the worksheet to a DataFrame
data = pd.DataFrame(ws.values)

# Skip the first row (header) and assign column names
data = data.iloc[1:]
data.columns = ['loss', 'accurate', 'id']

# Iterate through the rows, sum the values in columns loss, accurate, id, and store the result in a new column
for i in range(1, len(data)):
    data.loc[i, 'token'] = (7*data.loc[i-1, 'loss'] + 15*data.loc[i-1, 'accurate'])/162

# Remove the existing sheet
wb.remove(ws)

# Create a new sheet with the updated DataFrame
new_ws = wb.create_sheet(ws.title)
for r in pd.DataFrame(data.values).iterrows():
    for c in range(len(r[1])):
        new_ws.cell(row=r[0]+1, column=c+1, value=r[1][c])

# Save the updated workbook
wb.save(file_path)
